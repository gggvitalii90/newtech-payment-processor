from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from payment_processor.env import load_env
from payment_processor.fintablo_client import FinTabloClient, load_fintablo_settings
from payment_processor.google_api import build_sheets_service, get_credentials, load_google_settings
from scripts.fintablo_sync_from_manual_final import ManualRow, _u, norm, normalize_date, parse_amount

CASH_PAYMENT_TYPE = _u(r"\u041d\u0430\u043b\u0438\u0447\u043d\u0430\u044f")
KEY_DATE = _u(r"\u0414\u0430\u0442\u0430")
KEY_AMOUNT = _u(r"\u0421\u0443\u043c\u043c\u0430")
KEY_OPERATION_TYPE = _u(r"\u0422\u0438\u043f \u043e\u043f\u0435\u0440\u0430\u0446\u0438\u0438")
KEY_PAYMENT_TYPE = _u(r"\u0422\u0438\u043f \u043e\u043f\u043b\u0430\u0442\u044b")
KEY_OBJECT = _u(r"\u041e\u0431\u044a\u0435\u043a\u0442")
KEY_PROJECT = _u(r"\u041f\u0440\u043e\u0435\u043a\u0442")
KEY_BUDGET = _u(r"\u0421\u0442\u0430\u0442\u044c\u044f \u0431\u044e\u0434\u0436\u0435\u0442\u0430")
KEY_RESPONSIBLE = _u(r"\u041e\u0442\u0432\u0435\u0442\u0441\u0442\u0432\u0435\u043d\u043d\u044b\u0439")
KEY_PURPOSE = _u(r"\u041d\u0430\u0437\u043d\u0430\u0447\u0435\u043d\u0438\u0435 \u043f\u043b\u0430\u0442\u0435\u0436\u0430")
OPERATION_INCOME = _u(r"\u041f\u0440\u0438\u0445\u043e\u0434")
OPERATION_OUTCOME = _u(r"\u0420\u0430\u0441\u0445\u043e\u0434")
OPERATION_CONVERSION = _u(r"\u041a\u043e\u043d\u0432\u0435\u0440\u0442\u0430\u0446\u0438\u044f")


@dataclass
class CashReconciliationResult:
    summary: dict[str, Any]
    matched: list[dict[str, Any]]
    missing: list[dict[str, Any]]
    fintablo_unmatched: list[dict[str, Any]]


def parse_day(value: str) -> date:
    text = str(value or "").strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Unsupported date: {value}")


def display_day(value: date) -> str:
    return value.strftime("%d.%m.%Y")


def read_manual_rows_dynamic(start: str, end: str) -> list[ManualRow]:
    env = load_env()
    google_settings = load_google_settings(env)
    sheets = build_sheets_service(get_credentials(google_settings))
    spreadsheet_id = env["GOOGLE_DICTIONARY_SPREADSHEET_ID"]
    start_day = parse_day(start)
    end_day = parse_day(end)
    metadata = sheets.spreadsheets().get(
        spreadsheetId=spreadsheet_id,
        fields="sheets(properties(title))",
    ).execute()
    rows: list[ManualRow] = []
    for item in metadata.get("sheets", []):
        title = str(item.get("properties", {}).get("title") or "")
        if not re.search(r"20\d\d", title):
            continue
        values = sheets.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id,
            range=f"'{title}'!A1:O5000",
        ).execute().get("values", [])
        if not values:
            continue
        headers = [str(value).strip() for value in values[0]]
        if KEY_DATE not in headers or KEY_AMOUNT not in headers or KEY_PAYMENT_TYPE not in headers:
            continue
        for row_number, row in enumerate(values[1:], start=2):
            data = {headers[i]: str(row[i]).strip() if i < len(row) else "" for i in range(len(headers))}
            if not data.get(KEY_DATE) or not data.get(KEY_AMOUNT):
                continue
            try:
                row_day = parse_day(data[KEY_DATE])
            except ValueError:
                continue
            if start_day <= row_day <= end_day:
                rows.append(ManualRow(row_number, title, data))
    return rows


def moneybag_id(tx: dict[str, Any]) -> int:
    try:
        return int(tx.get("moneybagId") or 0)
    except (TypeError, ValueError):
        return 0


def moneybag2_id(tx: dict[str, Any]) -> int:
    try:
        return int(tx.get("moneybag2Id") or 0)
    except (TypeError, ValueError):
        return 0


def cash_moneybag_ids(moneybags: dict[int, dict[str, Any]]) -> set[int]:
    return {item_id for item_id, item in moneybags.items() if str(item.get("type") or "").strip() == "nal"}


def is_cash_tx(tx: dict[str, Any], moneybags: dict[int, dict[str, Any]]) -> bool:
    cash_ids = cash_moneybag_ids(moneybags)
    return moneybag_id(tx) in cash_ids or moneybag2_id(tx) in cash_ids


def cash_manual_rows(manual_rows: list[ManualRow]) -> list[ManualRow]:
    cash_key = norm(CASH_PAYMENT_TYPE)
    return [row for row in manual_rows if norm(row.payment_type).startswith(cash_key)]


def operation_group_from_row(row: ManualRow) -> str:
    group = row.operation_group
    if group == "transfer" and row.amount != 0:
        return "income" if row.amount > 0 else "outcome"
    return group


def operation_group_from_tx(tx: dict[str, Any], moneybags: dict[int, dict[str, Any]] | None = None) -> str:
    group = str(tx.get("group") or "").strip()
    if group != "transfer":
        return group if group in {"income", "outcome"} else "outcome"
    if moneybags is None:
        return "transfer"
    cash_ids = cash_moneybag_ids(moneybags)
    if moneybag2_id(tx) in cash_ids:
        return "income"
    if moneybag_id(tx) in cash_ids:
        return "outcome"
    return "transfer"


def amount_key(value: Any) -> Decimal:
    return abs(parse_amount(value))


def signed_amount_for_group(group: str, value: Any) -> Decimal:
    parsed = amount_key(value)
    if group == "income":
        return parsed
    if group == "outcome":
        return -parsed
    return parse_amount(value)


def signed_amount_row(row: ManualRow) -> Decimal:
    return signed_amount_for_group(operation_group_from_row(row), row.amount)


def signed_amount_tx(tx: dict[str, Any], moneybags: dict[int, dict[str, Any]]) -> Decimal:
    return signed_amount_for_group(operation_group_from_tx(tx, moneybags), tx.get("value"))


def text_score(manual_text: str, tx_text: str) -> int:
    manual_key = norm(manual_text)
    tx_key = norm(tx_text)
    if not manual_key or not tx_key:
        return 0
    score = 0
    if manual_key == tx_key:
        score += 100
    elif manual_key in tx_key or tx_key in manual_key:
        score += 40
    manual_words = [word for word in manual_key.split() if len(word) >= 4]
    tx_words = set(tx_key.split())
    score += sum(3 for word in manual_words[:12] if word in tx_words)
    return score


def match_score(row: ManualRow, tx: dict[str, Any], moneybags: dict[int, dict[str, Any]]) -> int:
    if normalize_date(tx.get("date", "")) != row.date:
        return -1
    if amount_key(tx.get("value")) != amount_key(row.amount):
        return -1
    if operation_group_from_row(row) != operation_group_from_tx(tx, moneybags):
        return -1
    return 10 + text_score(row.values.get(KEY_PURPOSE, ""), str(tx.get("description") or ""))


def row_payload(row: ManualRow) -> dict[str, Any]:
    values = row.values
    return {
        "manual_sheet": row.sheet,
        "manual_row": row.row_number,
        "date": row.date,
        "operation_type": values.get(KEY_OPERATION_TYPE, ""),
        "payment_type": values.get(KEY_PAYMENT_TYPE, ""),
        "amount": str(row.amount),
        "object": values.get(KEY_OBJECT, ""),
        "project": values.get(KEY_PROJECT, ""),
        "budget": values.get(KEY_BUDGET, ""),
        "responsible": values.get(KEY_RESPONSIBLE, ""),
        "purpose": values.get(KEY_PURPOSE, ""),
    }


def tx_payload(tx: dict[str, Any], moneybags: dict[int, dict[str, Any]]) -> dict[str, Any]:
    moneybag = moneybags.get(moneybag_id(tx), {})
    moneybag2 = moneybags.get(moneybag2_id(tx), {})
    return {
        "fintablo_id": tx.get("id", ""),
        "fintablo_date": normalize_date(tx.get("date", "")),
        "fintablo_group": tx.get("group", ""),
        "fintablo_cash_group": operation_group_from_tx(tx, moneybags),
        "fintablo_amount": str(parse_amount(tx.get("value"))),
        "fintablo_moneybag": moneybag.get("name", ""),
        "fintablo_moneybag2": moneybag2.get("name", ""),
        "fintablo_description": tx.get("description", ""),
    }


def reconcile_cash(manual_rows: list[ManualRow], transactions: list[dict[str, Any]], moneybags: dict[int, dict[str, Any]]) -> CashReconciliationResult:
    manual_cash = cash_manual_rows(manual_rows)
    fintablo_cash = [tx for tx in transactions if is_cash_tx(tx, moneybags)]
    used_tx_indexes: set[int] = set()
    matched: list[dict[str, Any]] = []
    missing: list[dict[str, Any]] = []
    missing_rows: list[ManualRow] = []

    for row in sorted(manual_cash, key=lambda item: (parse_day(item.date), item.amount, item.row_number)):
        candidates = []
        for index, tx in enumerate(fintablo_cash):
            if index in used_tx_indexes:
                continue
            score = match_score(row, tx, moneybags)
            if score >= 10:
                candidates.append((score, index, tx))
        candidates.sort(key=lambda item: item[0], reverse=True)
        if candidates:
            score, index, tx = candidates[0]
            used_tx_indexes.add(index)
            matched.append({**row_payload(row), **tx_payload(tx, moneybags), "match_score": score})
        else:
            missing.append(row_payload(row))
            missing_rows.append(row)

    fintablo_unmatched = [
        tx_payload(tx, moneybags)
        for index, tx in enumerate(fintablo_cash)
        if index not in used_tx_indexes
    ]
    manual_total = sum((amount_key(row.amount) for row in manual_cash), Decimal("0.00"))
    fintablo_total = sum((amount_key(tx.get("value")) for tx in fintablo_cash), Decimal("0.00"))
    matched_total = sum((amount_key(row["amount"]) for row in matched), Decimal("0.00"))
    missing_total = sum((amount_key(row["amount"]) for row in missing), Decimal("0.00"))
    manual_signed_total = sum((signed_amount_row(row) for row in manual_cash), Decimal("0.00"))
    fintablo_signed_total = sum((signed_amount_tx(tx, moneybags) for tx in fintablo_cash), Decimal("0.00"))
    missing_signed_total = sum((signed_amount_row(row) for row in missing_rows), Decimal("0.00"))
    raw_groups = _count_by(fintablo_cash, lambda tx: str(tx.get("group") or ""))
    cash_groups = _count_by(fintablo_cash, lambda tx: operation_group_from_tx(tx, moneybags))
    manual_groups = _count_by(manual_cash, operation_group_from_row)
    summary = {
        "manual_cash_rows": len(manual_cash),
        "manual_income_rows": manual_groups.get("income", 0),
        "manual_outcome_rows": manual_groups.get("outcome", 0),
        "manual_transfer_rows": manual_groups.get("transfer", 0),
        "fintablo_cash_rows": len(fintablo_cash),
        "fintablo_raw_income_rows": raw_groups.get("income", 0),
        "fintablo_raw_outcome_rows": raw_groups.get("outcome", 0),
        "fintablo_raw_transfer_rows": raw_groups.get("transfer", 0),
        "fintablo_cash_income_rows": cash_groups.get("income", 0),
        "fintablo_cash_outcome_rows": cash_groups.get("outcome", 0),
        "matched": len(matched),
        "missing": len(missing),
        "fintablo_unmatched": len(fintablo_unmatched),
        "manual_cash_total_abs": str(manual_total),
        "fintablo_cash_total_abs": str(fintablo_total),
        "matched_total_abs": str(matched_total),
        "missing_total_abs": str(missing_total),
        "fintablo_unmatched_total_abs": str(fintablo_total - matched_total),
        "manual_cash_net": str(manual_signed_total),
        "fintablo_cash_net": str(fintablo_signed_total),
        "net_difference_manual_minus_fintablo": str(manual_signed_total - fintablo_signed_total),
        "missing_net": str(missing_signed_total),
    }
    return CashReconciliationResult(summary, matched, missing, fintablo_unmatched)


def _count_by(items: list[Any], key_func: Any) -> dict[str, int]:
    result: dict[str, int] = {}
    for item in items:
        key = str(key_func(item) or "")
        result[key] = result.get(key, 0) + 1
    return result


def write_sheet(wb: Workbook, title: str, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet(title)
    fields = sorted({key for row in rows for key in row}) if rows else ["empty"]
    ws.append(fields)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for row in rows:
        ws.append([row.get(field, "") for field in fields])
    for column_index, field in enumerate(fields, start=1):
        max_len = max([len(str(field))] + [len(str(row.get(field, ""))) for row in rows])
        ws.column_dimensions[get_column_letter(column_index)].width = min(max(max_len + 2, 12), 60)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_report(path: Path, result: CashReconciliationResult, *, start: str, end: str) -> None:
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "summary"
    summary_ws.append(["metric", "value"])
    for cell in summary_ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    summary_ws.append(["start", start])
    summary_ws.append(["end", end])
    for key, value in result.summary.items():
        summary_ws.append([key, value])
    summary_ws.column_dimensions["A"].width = 38
    summary_ws.column_dimensions["B"].width = 24
    write_sheet(wb, "missing_manual_cash", result.missing)
    write_sheet(wb, "matched_cash", result.matched)
    write_sheet(wb, "fintablo_unmatched_cash", result.fintablo_unmatched)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def load_moneybags_by_id(client: FinTabloClient) -> dict[int, dict[str, Any]]:
    result: dict[int, dict[str, Any]] = {}
    for item in client.list_moneybags():
        try:
            result[int(item.get("id") or 0)] = item
        except (TypeError, ValueError):
            continue
    return result


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Reconcile manual cash rows with FinTablo cash transactions")
    parser.add_argument("--start", default="01.01.2026", help="Start date YYYY-MM-DD or DD.MM.YYYY")
    parser.add_argument("--end", default=display_day(date.today()), help="End date YYYY-MM-DD or DD.MM.YYYY")
    parser.add_argument("--output", default="", help="XLSX report path")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    start = display_day(parse_day(args.start))
    end = display_day(parse_day(args.end))
    if parse_day(end) < parse_day(start):
        raise SystemExit("--end must be >= --start")
    output = Path(args.output) if args.output else Path("reports") / f"fintablo_cash_manual_reconciliation_{start.replace('.', '-')}_{end.replace('.', '-')}.xlsx"

    env = load_env()
    client = FinTabloClient(load_fintablo_settings(env))
    manual_rows = read_manual_rows_dynamic(start, end)
    transactions = client.list_transactions(date_from=start, date_to=end)
    moneybags = load_moneybags_by_id(client)
    result = reconcile_cash(manual_rows, transactions, moneybags)
    write_report(output, result, start=start, end=end)
    print(json.dumps({"start": start, "end": end, **result.summary, "output": str(output)}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
