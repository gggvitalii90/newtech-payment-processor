from __future__ import annotations

import argparse
import json
import sys
import time
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from payment_processor.dictionaries import normalize_key
from payment_processor.env import load_env
from payment_processor.fintablo_client import FinTabloClient, FinTabloError, load_fintablo_settings


def _u(*codes: int) -> str:
    return "".join(chr(code) for code in codes)


WORD_CREATE = _u(0x0441, 0x043e, 0x0437, 0x0434, 0x0430, 0x0442, 0x044c)
WORD_DELETE = _u(0x0443, 0x0434, 0x0430, 0x043b, 0x0438, 0x0442, 0x044c)
WORD_RENAME = _u(0x043f, 0x0435, 0x0440, 0x0435, 0x0438, 0x043c, 0x0435, 0x043d, 0x043e, 0x0432, 0x0430, 0x0442, 0x044c)
WORD_SKIP = _u(0x043f, 0x0440, 0x043e, 0x043f, 0x0443, 0x0441, 0x0442, 0x0438, 0x0442, 0x044c)
PHRASE_DO_NOT_CREATE = _u(0x043d, 0x0435, 0x20, 0x0441, 0x043e, 0x0437, 0x0434, 0x0430, 0x0432, 0x0430, 0x0442, 0x044c)
PHRASE_DO_NOT_TOUCH = _u(0x043d, 0x0435, 0x20, 0x0442, 0x0440, 0x043e, 0x0433, 0x0430, 0x0442, 0x044c)

DEFAULT_REVIEW = ROOT / "outputs" / "fintablo_reference_review" / "fintablo_reference_review_current.xlsx"
DEFAULT_REPORT = ROOT / "outputs" / "fintablo_reference_review" / "fintablo_reference_apply_latest.json"


def main() -> int:
    parser = argparse.ArgumentParser(description="Apply approved FinTablo reference review actions")
    parser.add_argument("--review", default=str(DEFAULT_REVIEW))
    parser.add_argument("--output", default=str(DEFAULT_REPORT))
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()

    review_path = Path(args.review)
    if not review_path.exists():
        print(json.dumps({"error": f"review file not found: {review_path}"}, ensure_ascii=False), file=sys.stderr)
        return 2

    client = FinTabloClient(load_fintablo_settings(load_env()))
    categories = client.list_categories()
    category_by_key = _by_key(categories)
    children = _children_by_parent(categories)
    report: dict[str, Any] = {"mode": "apply" if args.apply else "dry-run", "review": str(review_path), "operations": []}

    create_rows = _read_rows(review_path, sheet_index=1, name_col=2)
    for row in create_rows:
        action = _create_action(row)
        if action == "skip":
            report["operations"].append({"type": "create_category", **row, "status": "skipped", "reason": "review_skip"})
            continue
        if action != "create":
            report["operations"].append({"type": "create_category", **row, "status": "skipped", "reason": "review_no_action"})
            continue
        key = normalize_key(row["name"])
        if key in category_by_key:
            report["operations"].append({"type": "create_category", **row, "status": "skipped", "reason": "already_exists", "id": category_by_key[key].get("id")})
            continue
        payload = {"name": row["name"], "group": "outcome", "type": "operating", "description": "Created from Google reference sync"}
        op = {"type": "create_category", **row, "payload": payload}
        if args.apply:
            try:
                response = client.create_category(payload)
                op.update({"status": "applied", "items": response.items})
                if response.items:
                    category_by_key[normalize_key(str(response.items[0].get("name") or row["name"]))] = response.items[0]
                time.sleep(0.15)
            except FinTabloError as exc:
                op.update({"status": "failed", "error": str(exc)})
        else:
            op.update({"status": "planned"})
        report["operations"].append(op)

    delete_rows = _read_rows(review_path, sheet_index=2, name_col=2)
    for row in delete_rows:
        action = _delete_action(row)
        if action != "delete":
            report["operations"].append({"type": "delete_category", **row, "status": "skipped", "reason": "review_skip"})
            continue
        item = category_by_key.get(normalize_key(row["name"]))
        op = {"type": "delete_category", **row}
        if not item:
            op.update({"status": "skipped", "reason": "not_found"})
        elif item.get("isBuiltIn"):
            op.update({"status": "skipped", "reason": "built_in", "id": item.get("id")})
        elif children.get(int(item.get("id"))):
            op.update({"status": "skipped", "reason": "has_children", "id": item.get("id"), "children": [child.get("name") for child in children[int(item.get("id"))]]})
        elif args.apply:
            try:
                client.delete_category(item.get("id"))
                op.update({"status": "applied", "id": item.get("id")})
                time.sleep(0.15)
            except FinTabloError as exc:
                op.update({"status": "failed", "id": item.get("id"), "error": str(exc)})
        else:
            op.update({"status": "planned", "id": item.get("id")})
        report["operations"].append(op)

    summary = Counter(f"{op['type']}:{op['status']}" for op in report["operations"])
    report["summary"] = dict(summary)
    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"mode": report["mode"], "summary": report["summary"], "output": str(out)}, ensure_ascii=False, indent=2))
    return 0


def _read_rows(path: Path, *, sheet_index: int, name_col: int) -> list[dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[sheet_index]
    rows: list[dict[str, str]] = []
    for idx, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        values = tuple(values or ())
        name = _cell(values, name_col)
        if not name:
            continue
        rows.append({
            "sheet": ws.title,
            "row": idx,
            "raw_action": _cell(values, 1),
            "name": name,
            "target": _cell(values, 3),
            "comment": _cell(values, 3),
        })
    return rows


def _cell(values: tuple[Any, ...], one_based: int) -> str:
    idx = one_based - 1
    if idx >= len(values):
        return ""
    return str(values[idx] or "").strip()


def _create_action(row: dict[str, str]) -> str:
    text = normalize_key(f"{row.get('raw_action', '')} {row.get('comment', '')}")
    if PHRASE_DO_NOT_CREATE in text or PHRASE_DO_NOT_TOUCH in text or WORD_SKIP in text:
        return "skip"
    if WORD_CREATE in text:
        return "create"
    return "review"


def _delete_action(row: dict[str, str]) -> str:
    text = normalize_key(f"{row.get('raw_action', '')} {row.get('comment', '')}")
    if WORD_DELETE in text:
        return "delete"
    if WORD_RENAME in text:
        return "review"
    return "skip"


def _by_key(items: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    for item in items:
        name = str(item.get("name") or "").strip()
        if not name:
            continue
        result.setdefault(normalize_key(name), item)
    return result


def _children_by_parent(items: list[dict[str, Any]]) -> dict[int, list[dict[str, Any]]]:
    children: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for item in items:
        parent_id = item.get("parentId")
        if parent_id:
            children[int(parent_id)].append(item)
    return children


if __name__ == "__main__":
    raise SystemExit(main())
