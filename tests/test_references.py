from openpyxl import Workbook, load_workbook

from payment_processor.models import PaymentRecord
from payment_processor.references import load_reference_lists
from payment_processor.workflow import write_records_to_workbook


def test_loads_reference_lists_from_workbook(tmp_path):
    reference = tmp_path / "reference.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([])
    sheet.append(["Банк", "Объект"])
    sheet.append(["б/н ИП Мочалов Сбер", "ПСК Ньютек"])
    workbook.save(reference)

    references = load_reference_lists(reference)

    assert "Банк" in references
    assert "б/н ИП Мочалов Сбер" in references["Банк"]
    assert "Объект" in references
    assert "ПСК Ньютек" in references["Объект"]


def test_workbook_gets_hidden_reference_sheet_and_dropdowns(tmp_path):
    output = tmp_path / "result.xlsx"
    record = PaymentRecord(
        name="sample.pdf",
        date="2026-06-09",
        operation_type="Расход",
        payment_type="Безналичные с НДС",
        bank="б/н Альфа",
        counterparty='ООО "ТЕСТ"',
        invoice_number="1",
        object_name="ПСК Ньютек",
        project="Офис",
        budget_item="Топливо",
        responsible="Родин.К",
        purpose="Оплата",
        invoice_link="",
        amount="1000",
    )
    references = {"Банк": ["б/н Альфа", "б/н Точка"], "Объект": ["ПСК Ньютек"]}

    write_records_to_workbook(output, "ПСК", [record], references)

    workbook = load_workbook(output)
    assert workbook["_Справочник"].sheet_state == "hidden"
    validations = list(workbook["ПСК"].data_validations.dataValidation)
    assert any(validation.type == "list" for validation in validations)
