from pathlib import Path

from openpyxl import load_workbook

from payment_processor.models import COLUMNS, PaymentRecord
from payment_processor.workflow import process_folder, write_records_to_workbook


ROOT = Path(__file__).resolve().parents[1] / ".staging" / "payment-history" / "payments"


def test_process_folder_returns_one_record_per_pdf():
    records = process_folder(ROOT / "2026.06.08", rules={})

    assert len(records) == 20
    assert {record.name for record in records if record.bank == "б/н Сбербанк"}
    assert {record.name for record in records if record.bank == "б/н Альфа"}


def test_process_folder_keeps_payment_orders_with_invoice_text_in_purpose():
    records = process_folder(ROOT / "2026.06.15", rules={})
    names = {record.name for record in records}

    assert len(records) == 21
    assert {
        "Платежное_поручение_№154.pdf",
        "Платежное_поручение_№155.pdf",
        "Платежное_поручение_№214.pdf",
        "Платежное_поручение_№215.pdf",
        "Платежное_поручение_№218.pdf",
        "Платежное_поручение_№219.pdf",
        "Платежное_поручение_№222.pdf",
        "Платежное_поручение_№223.pdf",
        "Платежное_поручение_№224.pdf",
        "Платежное_поручение_№225.pdf",
        "Платежное_поручение_№226.pdf",
    }.issubset(names)


def test_write_records_replaces_selected_sheet_and_keeps_other_sheet(tmp_path):
    output = tmp_path / "result.xlsx"
    psk_record = PaymentRecord(
        name="psk.pdf",
        date="2026-06-08",
        operation_type="Расход",
        payment_type="Безналичные с НДС",
        bank="б/н Альфа",
        counterparty='ООО "ТЕСТ"',
        invoice_number="1",
        object_name="",
        project="",
        budget_item="",
        responsible="",
        purpose="Оплата тест",
        invoice_link="",
        amount="1000",
    )
    is_record = PaymentRecord(
        name="is.pdf",
        date="2026-03-06",
        operation_type="Расход",
        payment_type="Безналичные с НДС",
        bank="б/н ИНВЕСТСТРОЙ",
        counterparty='ООО "ИС"',
        invoice_number="2",
        object_name="",
        project="",
        budget_item="",
        responsible="",
        purpose="Оплата ИС",
        invoice_link="",
        amount="2000",
    )

    write_records_to_workbook(output, "ИС", [is_record])
    write_records_to_workbook(output, "ПСК", [psk_record])

    workbook = load_workbook(output)
    assert workbook["ПСК"][1][0].value == COLUMNS[0]
    assert workbook["ПСК"][2][0].value == "psk.pdf"
    assert workbook["ИС"][2][0].value == "is.pdf"

    write_records_to_workbook(output, "ПСК", [])
    workbook = load_workbook(output)
    assert workbook["ПСК"].max_row == 1
    assert workbook["ИС"][2][0].value == "is.pdf"
