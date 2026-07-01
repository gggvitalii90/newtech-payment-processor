from pathlib import Path

from openpyxl import load_workbook


REFERENCE_COLUMNS = {
    "Тип операции",
    "Тип оплаты",
    "Банк",
    "Объект",
    "Проект",
    "Статья бюджета",
    "Ответственный",
}


def load_reference_lists(path: Path | str | None) -> dict[str, list[str]]:
    if not path:
        return {}
    path = Path(path)
    if not path.exists():
        return {}

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[2]]
    result: dict[str, list[str]] = {}

    for column_idx, header in enumerate(headers, start=1):
        if header not in REFERENCE_COLUMNS:
            continue
        values = result.setdefault(header, [])
        for row_idx in range(3, sheet.max_row + 1):
            value = sheet.cell(row_idx, column_idx).value
            if value is None:
                continue
            text = str(value).strip()
            if not text or text in values:
                continue
            values.append(text)
    return result

