from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

from .models import COLUMNS, PaymentRecord
from .parser import parse_payment_pdf
from .payment_classifier import is_payment_order_pdf


def process_folder(folder: Path, rules: dict | None = None) -> list[PaymentRecord]:
    pdfs = sorted(
        [
            path
            for path in folder.iterdir()
            if path.is_file() and path.suffix.lower() == ".pdf" and is_payment_order_pdf(path)
        ],
        key=lambda path: path.name.lower(),
    )
    return [parse_payment_pdf(path, rules or {}) for path in pdfs]


def write_records_to_workbook(
    output_path: Path,
    sheet_name: str,
    records: list[PaymentRecord],
    reference_lists: dict[str, list[str]] | None = None,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(output_path) if output_path.exists() else Workbook()
    if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) == 1 and workbook["Sheet"].max_row == 1:
        del workbook["Sheet"]
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    sheet = workbook.create_sheet(sheet_name)
    sheet.append(COLUMNS)
    for record in records:
        sheet.append(record.as_row())
    _format_sheet(sheet)
    if reference_lists:
        _write_reference_sheet(workbook, reference_lists)
        _add_validations(sheet, reference_lists)
    if workbook.sheetnames[0] != sheet_name and len(workbook.sheetnames) == 1:
        workbook.active = 0
    workbook.save(output_path)


def read_records_from_workbook(output_path: Path, sheet_name: str) -> list[PaymentRecord]:
    if not output_path.exists():
        return []
    workbook = load_workbook(output_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        return []
    sheet = workbook[sheet_name]
    records: list[PaymentRecord] = []
    for row in sheet.iter_rows(min_row=2, max_col=len(COLUMNS), values_only=True):
        if not any(value not in (None, "") for value in row):
            continue
        values = ["" if value is None else str(value) for value in row]
        records.append(PaymentRecord.from_row(values))
    return records


def _format_sheet(sheet) -> None:
    sheet.freeze_panes = "A2"
    widths = {
        "A": 34,
        "B": 14,
        "C": 16,
        "D": 22,
        "E": 18,
        "F": 38,
        "G": 18,
        "H": 18,
        "I": 20,
        "J": 22,
        "K": 18,
        "L": 80,
        "M": 18,
        "N": 14,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.auto_filter.ref = sheet.dimensions


def _write_reference_sheet(workbook, reference_lists: dict[str, list[str]]) -> None:
    if "_Справочник" in workbook.sheetnames:
        del workbook["_Справочник"]
    sheet = workbook.create_sheet("_Справочник")
    for col_idx, (header, values) in enumerate(reference_lists.items(), start=1):
        sheet.cell(row=1, column=col_idx, value=header)
        for row_idx, value in enumerate(values, start=2):
            sheet.cell(row=row_idx, column=col_idx, value=value)
    sheet.sheet_state = "hidden"


def _add_validations(sheet, reference_lists: dict[str, list[str]]) -> None:
    if not reference_lists:
        return
    ref_index = {header: idx for idx, header in enumerate(reference_lists, start=1)}
    for col_idx, header in enumerate(COLUMNS, start=1):
        values = reference_lists.get(header)
        if not values:
            continue
        ref_col = ref_index[header]
        col_letter = _column_letter(ref_col)
        formula = f"'_Справочник'!${col_letter}$2:${col_letter}${len(values) + 1}"
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        validation.error = "Выберите значение из справочника или оставьте ячейку пустой."
        validation.errorTitle = "Значение не из справочника"
        sheet.add_data_validation(validation)
        excel_col = _column_letter(col_idx)
        validation.add(f"{excel_col}2:{excel_col}1000")


def _column_letter(index: int) -> str:
    result = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        result = chr(65 + remainder) + result
    return result
